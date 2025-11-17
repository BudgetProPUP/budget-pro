import csv
import json
from decimal import Decimal

import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

from django.db import transaction
from django.db.models import Sum, Q, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mail
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import generics, status, viewsets, filters
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes, action
from drf_spectacular.utils import (
    extend_schema, OpenApiParameter, OpenApiResponse, OpenApiTypes, OpenApiExample
)

from .models import (
    Account, AccountType, BudgetProposal, Department, ExpenseCategory,
    FiscalYear, BudgetAllocation, Expense, JournalEntry, JournalEntryLine,
    ProposalComment, ProposalHistory, UserActivityLog, Project
)
from .permissions import IsTrustedService, IsBMSFinanceHead, IsBMSUser, IsBMSAdmin
from .pagination import FiveResultsSetPagination, SixResultsSetPagination, StandardResultsSetPagination
from .serializers import FiscalYearSerializer
from .serializers_budget import (
    AccountDropdownSerializer,
    AccountSetupSerializer,
    AccountTypeDropdownSerializer,
    BudgetAdjustmentSerializer,
    BudgetProposalListSerializer,
    BudgetProposalMessageSerializer,
    ProposalCommentCreateSerializer,
    BudgetProposalSummarySerializer,
    BudgetProposalDetailSerializer,
    DepartmentDropdownSerializer,
    ExpenseCategoryVarianceSerializer,
    JournalEntryCreateSerializer,
    JournalEntryListSerializer,
    LedgerViewSerializer,
    ProposalCommentSerializer,
    ProposalHistorySerializer,
    ProposalReviewBudgetOverviewSerializer,
    ProposalReviewSerializer
)


# --- Budget Proposal Page Views ---
# @extend_schema(
#     tags=['Budget Proposal Page'],
#     summary="List budget proposals with filters",
#     parameters=[
#         OpenApiParameter(name="department", type=str), OpenApiParameter(
#             name="status", type=str),
#         OpenApiParameter(name="search", type=str),
#     ],
#     responses={200: BudgetProposalListSerializer(many=True)}
# )
# class BudgetProposalListView(generics.ListAPIView):
#     serializer_class = BudgetProposalListSerializer
#     permission_classes = [IsAuthenticated]
#     # MODIFIED: Use new pagination class for 5 items per page
#     pagination_class = FiveResultsSetPagination

#     def get_queryset(self):  # Logic seems fine
#         queryset = BudgetProposal.objects.select_related('department').filter(
#             is_deleted=False)  # Added select_related and is_deleted filter
#         department_code = self.request.query_params.get('department')
#         status_filter = self.request.query_params.get('status')
#         search = self.request.query_params.get('search')
#         if department_code:
#             queryset = queryset.filter(department__code=department_code)
#         if status_filter:
#             queryset = queryset.filter(status=status_filter)
#         if search:
#             queryset = queryset.filter(
#                 Q(title__icontains=search) | Q(
#                     external_system_id__icontains=search)
#             )
#         return queryset.order_by('-created_at')  # Added default ordering


class BudgetProposalSummaryView(generics.GenericAPIView):
    # --- MODIFICATION START ---
    # Changed permission from IsAuthenticated to the more specific IsBMSUser
    permission_classes = [IsBMSUser]
    # --- MODIFICATION END ---
    serializer_class = BudgetProposalSummarySerializer

    @extend_schema(
        tags=['Budget Proposal Page'],
        summary="Get summary of budget proposals (cards)",
        responses={200: BudgetProposalSummarySerializer}
    )
    def get(self, request):
        user = request.user
        user_roles = getattr(user, 'roles', {})
        bms_role = user_roles.get('bms')

        # --- MODIFICATION START ---
        # Applied data isolation to the queryset used for the summary
        active_proposals = BudgetProposal.objects.filter(is_deleted=False)

        if bms_role == 'FINANCE_HEAD':
            department_id = getattr(user, 'department_id', None)
            if department_id:
                active_proposals = active_proposals.filter(department_id=department_id)
            else:
                active_proposals = BudgetProposal.objects.none()
        # Admins will see all proposals by default
        total = active_proposals.count()
        # Assuming PENDING means SUBMITTED for approval
        pending = active_proposals.filter(status='SUBMITTED').count()

        # Total budget from items of *approved* proposals if that's the intent for "total budget card"
        # Or all non-rejected proposals. Clarify business rule for "total budget" card.
        # For now, using all non-deleted proposals as per original:
        total_budget = active_proposals.aggregate(
            total=Sum('items__estimated_cost'))['total'] or 0
        data = {'total_proposals': total,
                'pending_approvals': pending, 'total_budget': total_budget}
        serializer = BudgetProposalSummarySerializer(
            data)  # Serialize the data
        return Response(serializer.data)


# class BudgetProposalDetailView(generics.RetrieveAPIView):  
#     queryset = BudgetProposal.objects.filter(is_deleted=False).prefetch_related('items__account', 'comments')
#     serializer_class = BudgetProposalDetailSerializer
#     permission_classes = [IsBMSUser] # JWT permission

#     @extend_schema(
#         tags=['Budget Proposal Page'],
#         summary="Retrieve full details of a proposal",
#         description="Returns details including items and estimated costs.",
#         responses={200: BudgetProposalDetailSerializer}
#     )
#     def get(self, request, *args, **kwargs):
#         return super().get(request, *args, **kwargs)


# --- Proposal History Page View ---
@extend_schema(
    tags=['Proposal History Page'],
    # Changed from "List proposals history"
    summary="List proposal history entries",
    parameters=[
        OpenApiParameter(name="search", type=str), OpenApiParameter(
            name="department", type=int),
        # This status refers to ProposalHistory.action or Proposal.status?
        OpenApiParameter(name="status", type=str),
        # Assuming it filters ProposalHistory by its 'action' field.
    ],
    responses={200: ProposalHistorySerializer(many=True)}
)
# This view lists ProposalHistory entries
class ProposalHistoryView(generics.ListAPIView):
    serializer_class = ProposalHistorySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = SixResultsSetPagination

    def get_queryset(self):
        qs = ProposalHistory.objects.select_related(
            'proposal__department'
        ).prefetch_related(
            'proposal__items__account__account_type'  # Add prefetch for better performance
        ).all()

        # Search by ticket ID or proposal title
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(proposal__title__icontains=search) | 
                Q(proposal__external_system_id__icontains=search)
            )

        # Filter by department
        department_id = self.request.query_params.get('department')
        if department_id:
            qs = qs.filter(proposal__department_id=department_id)

        # Filter by action/status (APPROVED, REJECTED, SUBMITTED, etc.)
        action_filter = self.request.query_params.get('action')  # Changed from 'status' to 'action'
        if action_filter:
            qs = qs.filter(action__iexact=action_filter)

        # Filter by category (AccountType: Asset, Expense, Liability, etc.)
        category = self.request.query_params.get('category')
        if category and category != "All Categories":
            # Filter proposals that have items with accounts of this type
            qs = qs.filter(
                proposal__items__account__account_type__name__iexact=category
            ).distinct()

        return qs.order_by('-action_at')

# --- Account Setup Page View ---


@extend_schema(
    tags=['Account Setup Page'],
    summary="List accounts with accomplishment status",
    parameters=[
        OpenApiParameter(name="fiscal_year_id", type=int, required=True),
        OpenApiParameter(name="search", type=str), OpenApiParameter(
            name="type", type=str),
        OpenApiParameter(name="status", type=str),
    ],
    responses={200: AccountSetupSerializer(many=True)}
)
# Logic seems fine, context for serializer is key
class AccountSetupListView(generics.ListAPIView):
    serializer_class = AccountSetupSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated]
    # queryset defined in get_queryset to ensure it's dynamic if needed

    def get_queryset(self):
        qs = Account.objects.select_related(
            'account_type').prefetch_related('allocations')
        search = self.request.query_params.get('search')
        # Changed from 'type' to avoid clash with built-in
        acc_type_name = self.request.query_params.get('type')
        status_filter = self.request.query_params.get('status')
        if search:
            qs = qs.filter(Q(code__icontains=search) |
                           Q(name__icontains=search))
        if acc_type_name:
            qs = qs.filter(account_type__name__iexact=acc_type_name)
        if status_filter:
            if status_filter.lower() == 'active':
                qs = qs.filter(is_active=True)
            elif status_filter.lower() == 'inactive':
                qs = qs.filter(is_active=False)
        return qs.order_by('code')

    def list(self, request, *args, **kwargs):  # Logic for passing fiscal_year context is fine
        fiscal_year_id = request.query_params.get('fiscal_year_id')
        if not fiscal_year_id:
            return Response({"error": "fiscal_year_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            fiscal_year = FiscalYear.objects.get(id=fiscal_year_id)
        except FiscalYear.DoesNotExist:
            return Response({"error": "Fiscal year not found"}, status=status.HTTP_404_NOT_FOUND)
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        # Pass request to context
        serializer_context = {'request': request, 'fiscal_year': fiscal_year}
        serializer = self.get_serializer(page, many=True, context=serializer_context) if page is not None else self.get_serializer(
            queryset, many=True, context=serializer_context)
        return self.get_paginated_response(serializer.data) if page is not None else Response(serializer.data)


@extend_schema(
    tags=['Fiscal Year Dropdown'],
    summary="For Fiscal Year values in a dropdown"
)
class FiscalYearDropdownView(generics.ListAPIView):
    queryset = FiscalYear.objects.all().order_by('-start_date')
    serializer_class = FiscalYearSerializer
    permission_classes = [IsAuthenticated]


@extend_schema(
    tags=['Ledger View'],
    summary="Get ledger view",
    parameters=[
            OpenApiParameter(
                name="search", description="Search by description or category", required=False, type=str),
            OpenApiParameter(
                name="category", description="Filter by category (e.g., EXPENSES)", required=False, type=str),
            OpenApiParameter(
                name="transaction_type", description="Filter by transaction type", required=False, type=str),
    ],
    responses={200: LedgerViewSerializer(many=True)}
)
# --- Ledger and Journal Entry Views ---
class LedgerViewList(generics.ListAPIView):
    serializer_class = LedgerViewSerializer
    # MODIFIED: Use new pagination class for 5 items per page
    pagination_class = FiveResultsSetPagination
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = JournalEntryLine.objects.select_related(
            'journal_entry', 'account')  # Added account
        search = self.request.query_params.get('search')
        category = self.request.query_params.get('category')
        transaction_type = self.request.query_params.get('transaction_type')
        if search:
            # MODIFICATION START: Add entry_id (Ticket ID) and date to the search fields
            queryset = queryset.filter(
                Q(journal_entry__entry_id__icontains=search) | # Search by Ticket ID
                Q(journal_entry__date__icontains=search) |       # Search by Date
                Q(journal_entry__description__icontains=search) | 
                Q(description__icontains=search) |
                Q(journal_entry__category__icontains=search) | 
                Q(account__name__icontains=search) | 
                Q(account__code__icontains=search)
            )
            # MODIFICATION END
        if category:
            queryset = queryset.filter(
                journal_entry__category__iexact=category)
        if transaction_type:
            queryset = queryset.filter(
                journal_transaction_type__iexact=transaction_type)
        # Added secondary sort
        return queryset.order_by('-journal_entry__date', 'journal_entry__entry_id')


@extend_schema(
    tags=['Ledger View'],
    summary="Export ledger entries as CSV",
    description="Exports filtered ledger entries in CSV format. Uses same filters as the ledger list view.",
    parameters=[
        OpenApiParameter(
            name="search", description="Search by description or category", required=False, type=str),
        OpenApiParameter(
            name="category", description="Filter by category (e.g., EXPENSES)", required=False, type=str),
        OpenApiParameter(name="transaction_type",
                         description="Filter by transaction type", required=False, type=str),
    ],
    responses={
        200: OpenApiResponse(
            description='CSV file attachment',
            response=OpenApiTypes.BINARY
        )
    }
)
class LedgerExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # This logic reuses the get_queryset from LedgerViewList for consistency
        list_view = LedgerViewList()
        list_view.request = request  # Mock the request for the view
        queryset = list_view.get_queryset()

        response = HttpResponse(
            content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="ledger_export.csv"'

        writer = csv.writer(response)
        # UPDATED: Match the new table columns
        writer.writerow(['Reference ID', 'Date', 'Category',
                        'Description', 'Account', 'Amount (PHP)'])

        for line in queryset:
            writer.writerow([
                line.journal_entry.entry_id,
                line.journal_entry.date.strftime('%Y-%m-%d'),
                line.journal_entry.category,
                line.description.replace('\n', ' ').strip(),
                line.account.name,
                "{:,.2f}".format(line.amount)
            ])
        return response

# Views of the Journal Entry page


@extend_schema(
    tags=['Journal Entry Page'],
    summary="List Journal Entries",
    parameters=[
            OpenApiParameter(
                "search", str, description="Search by description or reference"),
            OpenApiParameter(
                "category", str, description="Filter by category (e.g., EXPENSES, ASSETS)")
    ],
    responses={200: JournalEntryListSerializer(many=True)}
)
class JournalEntryListView(generics.ListAPIView):  # Logic seems fine
    serializer_class = JournalEntryListSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated]

    def get_queryset(self):  # Logic seems fine
        qs = JournalEntry.objects.all()
        search = self.request.query_params.get('search')
        category = self.request.query_params.get('category')
        if search:
            qs = qs.filter(Q(description__icontains=search)
                           | Q(entry_id__icontains=search))
        if category:
            qs = qs.filter(category__iexact=category)
        return qs.order_by('-date', '-entry_id')  # Added secondary sort

    def get(self, *args, **kwargs):
        return super().get(*args, **kwargs)


@extend_schema(
    tags=['Journal Entry Page'],
    summary="Create a Journal Entry. Make sure to also have two lines (DEBIT,CREDIT)",
    request=JournalEntryCreateSerializer,
    responses={201: OpenApiResponse(
        description="Journal entry created successfully")}
)
# View for Journal Entry Creating
# Serializer handles user context
class JournalEntryCreateView(generics.CreateAPIView):
    serializer_class = JournalEntryCreateSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


@extend_schema(
    tags=['Account Dropdown'],
    summary="List Active Accounts",
    description="Returns a list of active accounts for use in journal entry forms or dropdowns.",
    responses={200: OpenApiResponse(
        response=AccountDropdownSerializer(many=True))}
)
class AccountDropdownView(generics.ListAPIView):
    queryset = Account.objects.filter(
        is_active=True).select_related('account_type')
    serializer_class = AccountDropdownSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


@extend_schema(
    tags=['Journal/Ledger Category Dropdown'],
    methods=['GET'],
    summary="Get Journal Dropdown Choices",
    description="Returns valid options for journal entry categories, transaction types, and journal transaction types.",
    responses={
        200: OpenApiResponse(
            response={
                "categories": ["EXPENSES", "ASSETS", "PROJECTS", "VENDOR_CONTRACTS"],
                "transaction_types": ["DEBIT", "CREDIT"],
                "journal_transaction_types": ["CAPITAL_EXPENDITURE", "OPERATIONAL_EXPENDITURE", "TRANSFER"]
            }
        )
    }
)
@api_view(['GET'])  # journal_choices function - no change
@permission_classes([IsAuthenticated])
def journal_choices(request):
    # ... (same as your provided code)
    categories = [c[0]
                  for c in JournalEntry._meta.get_field('category').choices]
    transaction_types = [c[0] for c in JournalEntryLine._meta.get_field(
        'transaction_type').choices]
    journal_types = [c[0] for c in JournalEntryLine._meta.get_field(
        'journal_transaction_type').choices]
    return Response({"categories": categories, "transaction_types": transaction_types, "journal_transaction_types": journal_types})


@extend_schema(
    tags=['Department Dropdowns'],
    summary="List Departments",
    description="Returns a list of departments for use in filters or dropdowns.",
    responses={200: DepartmentDropdownSerializer(many=True)}
)
class DepartmentDropdownView(generics.ListAPIView):
    queryset = Department.objects.filter(is_active=True)
    serializer_class = DepartmentDropdownSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


@extend_schema(
    tags=['Account Dropdowns'],
    summary="List Account Types",
    description="Returns all account types for filtering or dropdowns.",
    responses={200: AccountTypeDropdownSerializer(many=True)}
)
class AccountTypeDropdownView(generics.ListAPIView):
    queryset = AccountType.objects.all()
    serializer_class = AccountTypeDropdownSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class ExternalBudgetProposalViewSet(viewsets.ModelViewSet):
    """
    ViewSet for EXTERNAL service-to-service communication (e.g., from DTS/TTS).
    Handles creating, updating, and deleting proposals via API Key authentication.
    """
    queryset = BudgetProposal.objects.filter(is_deleted=False)
    serializer_class = BudgetProposalMessageSerializer
    permission_classes = [IsTrustedService] # Enforces API Key for all actions

    # This ViewSet uses the default create(), update(), partial_update(), and destroy()
    # methods from ModelViewSet, which are now correctly protected.


class BudgetProposalUIViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for the User Interface (UI).
    Handles listing, retrieving, reviewing, and commenting on proposals for
    authenticated users with JWTs.
    """
    permission_classes = [IsBMSUser]
    pagination_class = FiveResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['title', 'external_system_id']
    filterset_fields = {
        'status': ['exact'],
        'items__account__account_type__id': ['exact'],
    }

    def get_queryset(self):
        """
        Dynamically filters the queryset based on user role for data isolation.
        - ADMINs see all proposals.
        - FINANCE_HEADs see proposals for their own department.
        """
        user = self.request.user
        # MODIFICATION START: Improve prefetch to include account_type for efficiency
        base_queryset = BudgetProposal.objects.filter(is_deleted=False).select_related(
            'department', 'fiscal_year'
        ).prefetch_related('items__account__account_type', 'comments')
        # MODIFICATION END

        user_roles = getattr(user, 'roles', {})
        bms_role = user_roles.get('bms')
        
        status_filter = self.request.query_params.get('status')
        department_filter = self.request.query_params.get('department')

        if bms_role == 'ADMIN':
            queryset = base_queryset
            if department_filter:
                queryset = queryset.filter(department__code=department_filter)
        elif bms_role == 'FINANCE_HEAD':
            department_id = getattr(user, 'department_id', None)
            if not department_id:
                return BudgetProposal.objects.none()
            queryset = base_queryset.filter(department_id=department_id)
        else:
            return BudgetProposal.objects.none()

        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset.order_by('-created_at')

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return BudgetProposalDetailSerializer
        # For 'list' action
        return BudgetProposalListSerializer

    @extend_schema(
        summary="Review a proposal (Finance Head)",
        request=ProposalReviewSerializer,
        responses={200: BudgetProposalDetailSerializer},
        tags=['Budget Proposal Page Actions']
    )
    @action(detail=True, methods=['post'], permission_classes=[IsBMSFinanceHead])
    def review(self, request, pk=None):
        proposal = self.get_object()
        review_input_serializer = ProposalReviewSerializer(data=request.data)
        review_input_serializer.is_valid(raise_exception=True)

        new_status = review_input_serializer.validated_data['status']
        comment_text = review_input_serializer.validated_data.get('comment', '')
        
        reviewer_user_id = request.user.id
        # Use get_full_name() for consistency
        reviewer_name = request.user.get_full_name() or request.user.username

        previous_status_for_history = proposal.status

        with transaction.atomic():
            proposal.status = new_status
            if new_status == 'APPROVED':
                proposal.approved_by_name = reviewer_name
                proposal.approval_date = timezone.now()
                if not hasattr(proposal, 'project') or proposal.project is None:
                    Project.objects.create(
                        name=f"Project for: {proposal.title}",
                        description=proposal.project_summary,
                        start_date=proposal.performance_start_date,
                        end_date=proposal.performance_end_date,
                        department=proposal.department,
                        budget_proposal=proposal,
                        status='PLANNING'
                    )
            elif new_status == 'REJECTED':
                proposal.rejected_by_name = reviewer_name
                proposal.rejection_date = timezone.now()
                if hasattr(proposal, 'project') and proposal.project is not None:
                    proposal.project.status = 'CANCELLED'
                    proposal.project.save(update_fields=['status'])

            proposal.last_modified = timezone.now()
            proposal.save()

            if comment_text:
                ProposalComment.objects.create(
                    proposal=proposal,
                    user_id=reviewer_user_id,
                    user_username=reviewer_name,
                    comment=comment_text
                )

            ProposalHistory.objects.create(
                proposal=proposal, action=new_status,
                action_by_name=reviewer_name,
                previous_status=previous_status_for_history,
                new_status=proposal.status,
                comments=comment_text or f"Status changed to {new_status}."
            )

        output_serializer = BudgetProposalDetailSerializer(proposal, context={'request': request})
        return Response(output_serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Add a comment to a proposal (internal user)",
        request=ProposalCommentCreateSerializer,
        responses={201: ProposalCommentSerializer},
        tags=['Budget Proposal Page Actions']
    )
    @action(detail=True, methods=['post'], url_path='add-comment', permission_classes=[IsBMSUser])
    def add_comment(self, request, pk=None):
        proposal = self.get_object()
        comment_serializer = ProposalCommentCreateSerializer(data=request.data)
        comment_serializer.is_valid(raise_exception=True)
        comment_text = comment_serializer.validated_data['comment']

        commenter_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
        
        comment_obj = ProposalComment.objects.create(
            proposal=proposal,
            user_id=request.user.id,
            user_username=commenter_name,
            comment=comment_text
        )
        ProposalHistory.objects.create(
            proposal=proposal, action='COMMENTED', action_by_name=commenter_name,
            comments=f"Comment added: '{comment_text}'"
        )
        return Response(ProposalCommentSerializer(comment_obj).data, status=status.HTTP_201_CREATED)
    
'''
Function that exports the budget proposal to an Excel File

Example:
const handleExportProposal = async (proposalId) => {
  try {
    const response = await fetch(
      `http://localhost:8000/api/budget-proposals/${proposalId}/export/`,
      {
        method: 'GET',
        headers: {
          // Add the Authorization header here
        },
      }
    );
'''


@extend_schema(
    tags=["Budget Proposal Export"],
    summary="Export a budget proposal to Excel",
    description="Exports the specified BudgetProposal and its items to an Excel file. Example:",
    responses={
        200: OpenApiResponse(description="XLSX file of the proposal"),
        404: OpenApiResponse(description="Proposal not found"),
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_budget_proposal_excel(request, proposal_id):
    try:
        proposal = BudgetProposal.objects.prefetch_related('items__account', 'department').get(
            id=proposal_id, is_deleted=False)  # Added is_deleted=False
    except BudgetProposal.DoesNotExist:
        return HttpResponse("Proposal not found", status=404)

    # --- START: Optional User Activity Logging ---
    try:
        UserActivityLog.objects.create(
            user_id=request.user.id,  # From JWT
            user_username=getattr(request.user, 'username', 'N/A'),  # From JWT
            log_type='EXPORT',  # Ensure 'EXPORT' is in UserActivityLog.LOG_TYPE_CHOICES
            action=f'Exported budget proposal to Excel: ID {proposal.id}, Title "{proposal.title}"',
            status='SUCCESS',
            details={
                'proposal_id': proposal.id,
                'proposal_title': proposal.title,
                'export_format': 'xlsx'
            }
        )
    except Exception as e:
        # Log an error if activity logging fails, but don't let it break the export
        print(f"Error logging export activity: {e}")
    # --- END: Optional User Activity Logging ---

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Proposal"

    # Header section
    header_labels = [
        "Title", "Project Summary", "Project Description",
        "Performance Start", "Performance End", "Performance Notes",
        "Department", "Submitted By", "Status"
    ]
    header_values = [
        proposal.title,
        proposal.project_summary,
        proposal.project_description,
        proposal.performance_start_date.strftime("%Y-%m-%d"),
        proposal.performance_end_date.strftime("%Y-%m-%d"),
        getattr(proposal, 'performance_notes', ''),
        proposal.department.name,
        proposal.submitted_by_name or "N/A",
        proposal.status
    ]
    for i, (label, value) in enumerate(zip(header_labels, header_values), start=1):
        ws.append([label, value])
        ws.cell(row=i, column=1).font = Font(
            bold=True)  # Bold the header label

    ws.append([])  # Blank row

    # able Header
    table_header = ["Account", "Cost Element",
                    "Description", "Estimated Cost", "Notes"]
    ws.append(table_header)
    table_header_row = ws.max_row
    for col in range(1, len(table_header) + 1):
        ws.cell(row=table_header_row, column=col).font = Font(bold=True)

    # Line Items
    total_cost = 0
    for item in proposal.items.all():
        ws.append([
            item.account.name if item.account else "N/A",
            item.cost_element,
            item.description,
            float(item.estimated_cost),
            item.notes or ""
        ])
        total_cost += item.estimated_cost

    # Total Row
    ws.append([])
    total_row = ws.max_row + 1
    ws.append(["", "", "Total", float(total_cost)])

    # Bold "Total", "Estimated Cost" header, and the summed value
    ws.cell(row=table_header_row, column=4).font = Font(
        bold=True)  # "Estimated Cost" header
    ws.cell(row=table_header_row, column=5).font = Font(
        bold=True)  # "Notes" header
    ws.cell(row=total_row, column=3).font = Font(
        bold=True)         # "Total" label
    ws.cell(row=total_row, column=4).font = Font(
        bold=True)         # summed value

    # Bold all "Description" column cells (header and values)
    for row in range(table_header_row, ws.max_row + 1):
        ws.cell(row=row, column=3).font = Font(bold=True)

    # Auto size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(
            col[0].column)].width = max(max_length + 2, 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"budget_proposal_{proposal.id}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


class BudgetVarianceReportView(APIView):
    permission_classes = [IsAuthenticated]
    '''
    View for budget variance report page
    '''
    @extend_schema(
        tags=["Budget Variance Reports"],
        summary="Budget Variance Report",
        description="The Budget Variance Report shows how budget allocations are utilized across hierarchical categories such as Income and Expense. For each category, it displays: Budget (total allocated from BudgetAllocation objects), Actual (total approved Expense entries), and Available (Budget minus Actual). A negative Available value indicates overspending. Categories are organized into levels: Level 1 (top-level like INCOME or EXPENSE), Level 2 (groups like DISCRETIONARY or OPERATIONS), and Level 3 (specific items like Cloud Hosting or Utilities). This report helps departments track financial performance and identify areas of over- or under-spending.",
        parameters=[
            OpenApiParameter(name="fiscal_year_id", required=True, type=int)
        ],
        responses={200: ExpenseCategoryVarianceSerializer(many=True)}
    )
    def get(self, request):
        fiscal_year_id = request.query_params.get('fiscal_year_id')
        if not fiscal_year_id:
            return Response({"error": "fiscal_year_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            fiscal_year = FiscalYear.objects.get(id=fiscal_year_id)
        except FiscalYear.DoesNotExist:
            return Response({"error": "Fiscal Year not found"}, status=status.HTTP_404_NOT_FOUND)

        # --- START: Optional User Activity Logging ---
        try:
            UserActivityLog.objects.create(
                user_id=request.user.id,  # From JWT
                user_username=getattr(
                    request.user, 'username', 'N/A'),  # From JWT
                log_type='REPORT_VIEW',  # Or 'REPORT_GENERATE', ensure type exists
                action=f'Generated Budget Variance Report for FY: {fiscal_year.name}',
                status='SUCCESS',
                details={
                    'fiscal_year_id': fiscal_year.id,
                    'fiscal_year_name': fiscal_year.name,
                    'report_type': 'BudgetVariance'
                }
            )
        except Exception as e:
            print(f"Error logging report generation activity: {e}")
        # --- END: Optional User Activity Logging ---

        top_categories = ExpenseCategory.objects.filter(
            level=1, is_active=True)

        def aggregate_node(category):
            # Direct values
            budget = BudgetAllocation.objects.filter(
                category=category, fiscal_year=fiscal_year, is_active=True
            ).aggregate(total=Coalesce(Sum('amount'), 0, output_field=DecimalField()))['total']

            actual = Expense.objects.filter(
                category=category,
                budget_allocation__fiscal_year=fiscal_year,
                status='APPROVED'
            ).aggregate(total=Coalesce(Sum('amount'), 0, output_field=DecimalField()))['total']

            available = budget - actual
            children = []

            for child in category.subcategories.all():
                children.append(aggregate_node(child))

            # Aggregate child totals if exists
            if children:
                child_budget = sum(c['budget'] for c in children)
                child_actual = sum(c['actual'] for c in children)
                child_available = child_budget - child_actual
                return {
                    "category": category.name,
                    "code": category.code,
                    "level": category.level,
                    "budget": round(child_budget, 2),
                    "actual": round(child_actual, 2),
                    "available": round(child_available, 2),
                    "children": children
                }
            else:
                return {
                    "category": category.name,
                    "code": category.code,
                    "level": category.level,
                    "budget": round(budget, 2),
                    "actual": round(actual, 2),
                    "available": round(available, 2),
                    "children": []
                }
        # Builds nested dictionariers, each representing top level expense categories, and its full budget/actual/available breakdown (and subcategories)
        # aggregate_node(cat) - recursive function that computes budget, actual, and available amounts for given category (and all its children if any)
        result = [aggregate_node(cat) for cat in top_categories]

        return Response(result)


@extend_schema(
    tags=["Budget Proposal Page Actions"],
    summary="Get budget overview for a proposal review",
    description="Provides financial context for a department's budget when reviewing a specific proposal.",
    responses={200: ProposalReviewBudgetOverviewSerializer}
)
class ProposalReviewBudgetOverview(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, proposal_id):
        try:
            proposal = BudgetProposal.objects.select_related(
                'department', 'fiscal_year').get(id=proposal_id)
        except BudgetProposal.DoesNotExist:
            return Response({"error": "Proposal not found"}, status=status.HTTP_404_NOT_FOUND)

        department = proposal.department
        fiscal_year = proposal.fiscal_year
        proposal_cost = proposal.items.aggregate(
            total=Coalesce(Sum('estimated_cost'), Decimal('0')))['total']

        # 1. Total budget allocated to the department for the fiscal year
        total_dept_budget = BudgetAllocation.objects.filter(
            department=department, fiscal_year=fiscal_year, is_active=True
        ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']

        # 2. Total spent by the department so far
        total_spent = Expense.objects.filter(
            department=department, budget_allocation__fiscal_year=fiscal_year, status='APPROVED'
        ).aggregate(total=Coalesce(Sum('amount'), Decimal('0')))['total']

        # UI's "Currently Allocated" seems to mean "already spent"
        currently_allocated = total_spent

        # 3. Available budget before considering this new proposal
        available_budget = total_dept_budget - total_spent

        # 4. Budget after this proposal is approved
        budget_after_proposal = available_budget - proposal_cost

        data = {
            "total_department_budget": total_dept_budget,
            "currently_allocated": currently_allocated,
            "available_budget": available_budget,
            "budget_after_proposal": budget_after_proposal
        }

        serializer = ProposalReviewBudgetOverviewSerializer(data)
        return Response(serializer.data)


@extend_schema(
    tags=["Budget Variance Reports"],
    summary="Export Budget Variance Report to Excel",
    description="Exports the hierarchical budget variance report to an XLSX file.",
    parameters=[
        OpenApiParameter(name="fiscal_year_id", required=True, type=int)
    ],
    responses={
        200: OpenApiResponse(description="XLSX file of the budget variance report"),
        400: OpenApiResponse(description="fiscal_year_id is required"),
        404: OpenApiResponse(description="Fiscal Year not found"),
    }
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_budget_variance_excel(request):
    fiscal_year_id = request.query_params.get('fiscal_year_id')
    if not fiscal_year_id:
        return Response({"error": "fiscal_year_id is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        fiscal_year = FiscalYear.objects.get(id=fiscal_year_id)
    except FiscalYear.DoesNotExist:
        return Response({"error": "Fiscal Year not found"}, status=status.HTTP_404_NOT_FOUND)

    # Re-use the logic from BudgetVarianceReportView to get the data
    top_categories = ExpenseCategory.objects.filter(level=1, is_active=True)

    def aggregate_node(category):
        budget = BudgetAllocation.objects.filter(
            category=category, fiscal_year=fiscal_year, is_active=True
        ).aggregate(total=Coalesce(Sum('amount'), 0, output_field=DecimalField()))['total']

        actual = Expense.objects.filter(
            category=category,
            budget_allocation__fiscal_year=fiscal_year,
            status='APPROVED'
        ).aggregate(total=Coalesce(Sum('amount'), 0, output_field=DecimalField()))['total']

        children_data = [aggregate_node(child)
                         for child in category.subcategories.all()]

        if children_data:
            child_budget = sum(c['budget'] for c in children_data)
            child_actual = sum(c['actual'] for c in children_data)
            return {
                "name": category.name, "budget": child_budget, "actual": child_actual,
                "available": child_budget - child_actual, "children": children_data
            }
        else:
            return {
                "name": category.name, "budget": budget, "actual": actual,
                "available": budget - actual, "children": []
            }

    report_data = [aggregate_node(cat) for cat in top_categories]

    # --- Excel Generation ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Variance Report"

    header = ["Category", "Budget", "Actual", "Available"]
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    def write_rows(data, indent=0):
        for item in data:
            ws.append([
                f"{' ' * indent * 4}{item['name']}",
                item['budget'],
                item['actual'],
                item['available']
            ])
            if item.get('children'):
                write_rows(item['children'], indent + 1)

    write_rows(report_data)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(
            col[0].column)].width = max(max_length + 2, 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"budget_variance_report_{fiscal_year.name}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@extend_schema(
    tags=["Budget Adjustment Page"],
    summary="Create a Budget Adjustment",
    description="Modifies a budget by transferring funds between allocations or adjusting against a reserve account, while creating a corresponding journal entry for audit purposes.",
    request=BudgetAdjustmentSerializer,
    # Returns the created journal entry
    responses={201: JournalEntryListSerializer}
)
class BudgetAdjustmentView(generics.CreateAPIView):
    """
    Handles the creation of a budget adjustment.
    - If both source and destination allocations are provided, it performs a transfer.
    - If only a source is provided, it adjusts the budget against an offsetting account.
    In both cases, it creates a balanced Journal Entry.
    """
    permission_classes = [IsAuthenticated]
    serializer_class = BudgetAdjustmentSerializer

    def perform_create(self, serializer):
        validated_data = serializer.validated_data
        user = self.request.user

        source_alloc = BudgetAllocation.objects.get(
            id=validated_data['source_allocation_id'])
        dest_alloc = None
        if validated_data.get('destination_allocation_id'):
            dest_alloc = BudgetAllocation.objects.get(
                id=validated_data['destination_allocation_id'])

        offsetting_account = Account.objects.get(
            id=validated_data['offsetting_account_id'])
        amount = validated_data['amount']
        description = validated_data['description']
        date = validated_data['date']

        with transaction.atomic():
            # Create the parent Journal Entry
            journal_entry = JournalEntry.objects.create(
                date=date,
                category='PROJECTS',  # Or a new 'ADJUSTMENTS' category
                description=description,
                total_amount=amount,
                status='POSTED',
                created_by_user_id=user.id,
                created_by_username=getattr(user, 'username', 'N/A')
            )

            if dest_alloc:  # This is a transfer between two allocations
                # 1. Decrease source allocation
                source_alloc.amount -= amount
                source_alloc.save(update_fields=['amount'])

                # 2. Increase destination allocation
                dest_alloc.amount += amount
                dest_alloc.save(update_fields=['amount'])

                # 3. Create Journal Entry Lines for the transfer
                # Credit the source allocation's account
                JournalEntryLine.objects.create(
                    journal_entry=journal_entry, account=source_alloc.account,
                    transaction_type='CREDIT', journal_transaction_type='TRANSFER',
                    amount=amount, description=f"Transfer from {source_alloc.project.name}"
                )
                # Debit the destination allocation's account
                JournalEntryLine.objects.create(
                    journal_entry=journal_entry, account=dest_alloc.account,
                    transaction_type='DEBIT', journal_transaction_type='TRANSFER',
                    amount=amount, description=f"Transfer to {dest_alloc.project.name}"
                )

            else:  # This is an adjustment against a single allocation
                # For this to be a balanced transaction, we need to decide if it's an increase or decrease
                # The UI modal's Debit/Credit can determine this. Let's assume the serializer will provide it.
                # For now, let's assume it's a decrease from the source allocation.

                # 1. Decrease source allocation
                source_alloc.amount -= amount
                source_alloc.save(update_fields=['amount'])

                # 2. Create Journal Entry Lines for the adjustment
                # Credit the source allocation's account (reducing its budget)
                JournalEntryLine.objects.create(
                    journal_entry=journal_entry, account=source_alloc.account,
                    transaction_type='CREDIT', journal_transaction_type='OPERATIONAL_EXPENDITURE',  # Or similar
                    amount=amount, description=f"Budget reduction for {source_alloc.project.name}"
                )
                # Debit the offsetting account (e.g., funds return to a general reserve)
                JournalEntryLine.objects.create(
                    journal_entry=journal_entry, account=offsetting_account,
                    transaction_type='DEBIT', journal_transaction_type='TRANSFER',
                    amount=amount, description="Return of funds to reserve"
                )

        # We return the created journal entry as proof of the transaction
        return_serializer = JournalEntryListSerializer(journal_entry)
        # The DRF CreateAPIView will wrap this in a 201 Response automatically.
        # To set the data for the response, we can set it on the instance.
        self.created_instance = journal_entry

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        # Use the serializer for the list view to format the response
        response_serializer = JournalEntryListSerializer(self.created_instance)
        headers = self.get_success_headers(response_serializer.data)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED, headers=headers)
